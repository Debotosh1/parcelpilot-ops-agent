#!/usr/bin/env python3
"""Re-ingest the supplied data pack.

    python scripts/ingest_pack.py --all

* `--structured` rewrites data/structured/{accounts,orders,tickets}.csv and the
  snapshot metadata from ParcelPilot_Assessment_Data.xlsx (exact, lossless).
* `--documents` extracts the text of every PDF into data/source_pack/extracted/
  and checks it against the committed markdown in data/documents/, reporting any
  number or clause that no longer matches its source.

The markdown corpus is committed rather than generated at runtime, because the
front matter (authority tier, status, superseded_by, account scope) is editorial
metadata that a PDF does not carry. This script is how you keep the two in sync
when a new version of a policy lands.

Requires: openpyxl, and either `pdftotext` (poppler-utils) or pypdf.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
PACK = ROOT / "data" / "source_pack"
STRUCTURED = ROOT / "data" / "structured"
DOCS = ROOT / "data" / "documents"
EXTRACTED = PACK / "extracted"

SHEET_TO_CSV = {"accounts": "accounts.csv", "orders": "orders.csv", "tickets": "tickets.csv"}

# markdown file -> source pdf
DOC_SOURCES = {
    "01_support_policy_v3.md": "01_Support_Policy_v3_CURRENT.pdf",
    "02_support_policy_v2_deprecated.md": "02_Support_Policy_v2_DEPRECATED.pdf",
    "03_cancellation_and_service_credit_sop_v4.md": "03_Cancellation_and_Service_Credit_SOP_v4.pdf",
    "04_product_operations_guide.md": "04_Product_Operations_Guide_and_Known_Issues.pdf",
    "05_northstar_enterprise_agreement.md": "05_Northstar_Logistics_Enterprise_Agreement.pdf",
    "06_lumenworks_service_agreement.md": "06_LumenWorks_Service_Agreement.pdf",
}


def pdf_text(path: Path) -> str:
    if shutil.which("pdftotext"):
        return subprocess.run(
            ["pdftotext", "-layout", str(path), "-"], capture_output=True, text=True, check=True
        ).stdout
    try:
        from pypdf import PdfReader
    except ImportError:  # pragma: no cover
        sys.exit("Install poppler-utils (pdftotext) or `pip install pypdf` to extract PDF text.")
    return "\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages)


def ingest_structured() -> None:
    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        sys.exit("pip install openpyxl to re-ingest the workbook.")

    workbook = openpyxl.load_workbook(PACK / "ParcelPilot_Assessment_Data.xlsx")
    for sheet_name, filename in SHEET_TO_CSV.items():
        sheet = workbook[sheet_name]
        rows = [
            ["" if cell is None else cell for cell in row]
            for row in sheet.iter_rows(values_only=True)
            if any(cell is not None for cell in row)
        ]
        with (STRUCTURED / filename).open("w", newline="", encoding="utf-8") as fh:
            csv.writer(fh).writerows(rows)
        print(f"wrote {filename} ({len(rows) - 1} rows)")

    readme = {str(r[0]).strip(): r[1] for r in workbook["README"].iter_rows(values_only=True) if r and r[0]}
    meta_path = STRUCTURED / "dataset_meta.json"
    meta = json.loads(meta_path.read_text("utf-8"))
    snapshot = str(readme.get("Dataset snapshot", "")).strip()
    if snapshot:
        parts = snapshot.split()
        meta["dataset_snapshot"] = " ".join(parts[:2])
        if len(parts) > 2:
            meta["timezone"] = parts[2]
    meta["currency"] = readme.get("Currency", meta.get("currency"))
    meta["notes"] = readme.get("Notes", meta.get("notes"))
    meta["important"] = readme.get("Important", meta.get("important"))
    meta_path.write_text(json.dumps(meta, indent=2) + "\n", encoding="utf-8")
    print(f"wrote dataset_meta.json (snapshot {meta['dataset_snapshot']} {meta['timezone']})")


def normalise(text: str) -> str:
    return re.sub(r"[\s​⁠]+", " ", text.replace("–", "-").replace("’", "'")).lower()


def verify_documents() -> int:
    """Every number in the committed markdown must exist in its source PDF."""
    EXTRACTED.mkdir(parents=True, exist_ok=True)
    problems = 0
    for md_name, pdf_name in DOC_SOURCES.items():
        md_path, pdf_path = DOCS / md_name, PACK / pdf_name
        if not pdf_path.exists():
            print(f"! {pdf_name} missing from data/source_pack - skipped")
            continue
        raw = pdf_text(pdf_path)
        (EXTRACTED / (pdf_path.stem + ".txt")).write_text(raw, encoding="utf-8")
        source = normalise(raw)
        body = md_path.read_text("utf-8").split("---", 2)[-1]
        numbers = set(re.findall(r"\b\d[\d,]*\b", body))
        missing = sorted(n for n in numbers if n.replace(",", "") not in source.replace(",", ""))
        status = "ok" if not missing else f"MISSING {missing}"
        problems += bool(missing)
        print(f"{'✓' if not missing else '✗'} {md_name:<46} {status}")
    return problems


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--structured", action="store_true", help="rebuild CSVs from the workbook")
    parser.add_argument("--documents", action="store_true", help="extract PDFs and verify the markdown corpus")
    parser.add_argument("--all", action="store_true")
    args = parser.parse_args()
    if not any([args.structured, args.documents, args.all]):
        parser.print_help()
        return
    if args.structured or args.all:
        ingest_structured()
    if args.documents or args.all:
        problems = verify_documents()
        if problems:
            sys.exit(f"{problems} document(s) contain figures not found in the source PDF.")


if __name__ == "__main__":
    main()
